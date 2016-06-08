import os
import re
import matplotlib.pyplot as plt
import openpyxl
workbook=openpyxl.Workbook()    #creates openpyxl workbook



'''columns for log file data worksheets'''
colMolecule = 'A'
colSym = 'B'
colCharge = 'C'
colMultiplicity = 'D'
colBasis = 'E'
colCCSDT = 'F'
colHF='G'
colCORR='H'
colMP2='I'
colMP3='J'
colMP4D='K'
colMP4DQ='L'
colMP4SDQ='M'
colCCSD = 'N'
colOrbital='O'
colElectronicState='P'

'''columns for VDE worksheet'''
colVDEMolecule='A'
colVDESymmetry='B'
colVDEBasis='C'
colVDE_HFcharged='D'
colVDE_HFneutral='E'
colVDE_HF_Ha='F'
colVDE_HF_eV='G'
colVDE_CCSDTcharged='H'
colVDE_CCSDTneutral='I'
colVDE_CCSDT_Ha='J'
colVDE_CCSDT_eV='K'
colVDE_CORRcharged='L'
colVDE_CORRneutral='M'
colVDE_CORR_Ha='N'
colVDE_CORR_eV='O'

'''columns for CORR worksheet'''
colCORRmolecule='A'
colCORRaugmented='B'
colCORRcharge='C'
colCORRbasisX='D'
colCORRbasisY='E'
colCORR_Ecorr='F'
colCORR_A='G'
colCORR_VDE_CORR='H'
colMaxBasisSet='I'
colVDEcc_hfcorr='J'
colVDEcc_eV = 'K'


'''path to this file'''
path=os.path.dirname(os.path.realpath(__file__))
pathorigin=path     #location to save workbook
#/Users/Jared/Dropbox/Auburn/Research/Second_Research/Log_Files

'''excel file name to open with path'''
excelFilePathName='/logFiles_openpyxl_test.xlsx'

'''folder containing the log files'''
logFilesFolder='/Extrapolation_Molecules'

chargedFolder='charged/'
neutralFolder='neutral/'
augFolder='aug/'
ccFolder='cc/'

'''list of the four basis sets in order increasing size'''
basisSets = ['pvDZ', 'pvTZ', 'pvQZ', 'pv5Z']

'''Ha to eV conversion factor'''
Ha_eV_conversion=27.21138602

'''sheet numbers in logFiles excel file'''
sheetNeutral=0
sheetCharged=1
sheetVDE=2
sheetCORR=3

neutralName="NEUTRAL"
chargedName="CHARGED"
vdeName="VDE"
corrName="CORR"


'''color for graphs'''
primaryColor=(.22, .42, .69)     #blue
secondaryColor=(1.0, .84, 0)    #gold
tertiaryColor=(0, 1, 1) #mystery

corrSheetRow=2
worksheetCORR=None

correlationMoleculeDict={}

vdeCCSDTev_list=[]
'''
list of dictionaries
(molecule, augmented, (basisSet/sets, corrValue))
[moleculeaug, molecule, moleculeaug, molecule]

each dictionary
basisSet/sets = Ecorr
'''
runOnce=False
timesRun=0

worksheetNeutral = workbook.active
worksheetNeutral.title=neutralName
worksheetCharged = workbook.create_sheet(title=chargedName)
worksheetVDE = workbook.create_sheet(title=vdeName)   #create VDE worksheet #3
worksheetCORR = workbook.create_sheet(title=corrName)   #create CORR worksheet #4

def run():
    """function calls dataExtract """
    dataExtract(path)

def rowsForSymAndMol(molecule, symmetry, basis, worksheetName):
   #returns the starting and ending rows in the excel file based off of the molecule and symmetry
    ##print(worksheetName)
    #if worksheetName==neutralName:
    #    worksheetNumber=sheetNeutral
    #if worksheetName==chargedName:
    #    worksheetNumber=sheetCharged
    #excelFile=pd.read_excel(path + excelFilePathName, sheetname=worksheetNumber)


    molStartRow=0   #first row with molecule and symmetry
    molEndRow=0     #last row with molecule and symmetry
    row=1   #current row in excel files
    endFound=False    #true when last row of molecule and symmetry is found
    startFound=False    #true when first row of molecule and symmetry is found
    while endFound==False:

        '''try:
            #str(excelFile.ix[row,0])
            str(worksheetName['A' + str(row)].value)
        except:
            molEndRow=row-1
            endFound=True
            break
        '''
        #if str(excelFile.ix[row,colMolecule]) == molecule and str(excelFile.ix[row,colSym]) \
        #== symmetry and str(excelFile.ix[row, colBasis][0:2])==basis[0:2]:
        if str(worksheetName[colMolecule + str(row)].value)==molecule and \
        str(worksheetName[colSym + str(row)].value)==symmetry and \
        str(worksheetName[colBasis + str(row)].value)[0:2]==basis[0:2]:
            if startFound==False:
                molStartRow=row
                startFound=True
        elif startFound==True and endFound==False:
            molEndRow=row-1
            endFound=True
        row+=1

        #print(molStartRow,molEndRow)

    print('molecule is ' + molecule)
    print('    mol start row is ' + str(molStartRow))
    print('    mol end round is ' + str(molEndRow))
    return(molStartRow,molEndRow)


def graph_OneLine(values, l, labels, molecule, worksheetName, augmented, yLabel, graphFolder):
    #graph_HF_CORR
    #line graphs HF values and corr values
    #s all graphs with just one line

        if worksheetName==chargedName:
            chargeFolder=chargedFolder
        if worksheetName==neutralName:
            chargeFolder=neutralFolder
        if augmented==True:
            augmentedFolder=augFolder
        if augmented==False:
            augmentedFolder=ccFolder

        plt.plot(l, values, color=primaryColor, lw=2, ls='-', marker='s', label=molecule)

        plt.xticks(l, labels, rotation = '30', ha='right')
        plt.margins(0.09, 0.09)

        #y_formatter = plt.ticker.ScalarFormatter(useOffset=False)
        #ax.yaxis.set_major_formatter(y_formatter)

        plt.subplots_adjust(bottom=0.2, top=0.85)
        plt.ylabel(yLabel)
        plt.legend(loc='upper center', bbox_to_anchor=(.5, 1.2), numpoints = 1, shadow=True, ncol=3)
        plt.grid(True)
        if not os.path.exists(path + '/ALL GRAPHS' + graphFolder +chargeFolder+augmentedFolder):
            os.makedirs(path + '/ALL GRAPHS' + graphFolder +chargeFolder+augmentedFolder)
        plt.savefig(path + '/ALL GRAPHS' + graphFolder +chargeFolder+augmentedFolder+ molecule + '.eps')
        #plt.show()
        plt.close()

def graph_TwoLines(aValues, bValues, l, labels, molecule, worksheetName, augmented, yLabel, graphFolder, aLabel, bLabel):
    #graph_HFandCCSDT
    '''graphs CCSDT and HF on same axis'''
    if worksheetName==chargedName:
        chargeFolder=chargedFolder
    if worksheetName==neutralName:
        chargeFolder=neutralFolder
    if augmented==True:
        augmentedFolder=augFolder
    if augmented==False:
        augmentedFolder=ccFolder

    plt.plot(l, aValues, color=primaryColor, lw=2, ls='-', marker='s', label=molecule + aLabel)
    plt.plot(l, bValues, color=secondaryColor, lw=2, ls='-', marker='o', label=molecule + bLabel)

    plt.xticks(l, labels, rotation = '30', ha='right')
    plt.margins(0.015, 0.05)
    plt.subplots_adjust(bottom=0.2, top=0.85)
    plt.ylabel(yLabel)
    plt.legend(loc='upper center', bbox_to_anchor=(.5, 1.2), numpoints = 1, shadow=True, ncol=2)
    plt.grid(True)

    if not os.path.exists(path + '/ALL GRAPHS' + graphFolder +chargeFolder+augmentedFolder):
            os.makedirs(path + '/ALL GRAPHS' + graphFolder +chargeFolder+augmentedFolder)
    plt.savefig(path + '/ALL GRAPHS' + graphFolder +chargeFolder+augmentedFolder+ molecule + '.eps')
    #plt.show()
    plt.close()


def graphVDE(values, l, labels, molecule, augmented, yLabel, graphFolder):
    '''line graphs VDEhf'''

    if augmented==True:
        augmentedFolder=augFolder
    if augmented==False:
        augmentedFolder=ccFolder

    plt.plot(l, values, color=primaryColor, lw=2, ls='-', marker='s', label=molecule)

    plt.xticks(l, labels, rotation = '30', ha='right')
    plt.margins(0.015, 0.05)
    plt.subplots_adjust(bottom=0.2, top=0.85)
    plt.ylabel(yLabel)
    plt.legend(loc='upper center', bbox_to_anchor=(.5, 1.2), numpoints = 1, shadow=True, ncol=3)
    plt.grid(True)
    if not os.path.exists(path+ '/ALL GRAPHS' + graphFolder+augmentedFolder):
            os.makedirs(path+ '/ALL GRAPHS' + graphFolder+augmentedFolder)
    plt.savefig(path+ '/ALL GRAPHS' + graphFolder+augmentedFolder + molecule + '.eps')
    #plt.show()
    plt.close()

def graph_VDEcombined_CCSDT_CORR_HF(hfValues, ccsdtValues, corrValues, l, labels, molecule, sheetNumber, augmented, \
    yLabel, graphFolder):
    if augmented==True:
        augmentedFolder=augFolder
    if augmented==False:
        augmentedFolder=ccFolder

    plt.plot(l, hfValues, color=primaryColor, lw=2, ls='-', marker='s', label=molecule + ' VDE HF')
    plt.plot(l, ccsdtValues, color=secondaryColor, lw=2, ls='-', marker='o', label=molecule + ' VDE CCSD(T)')
    plt.plot(l, corrValues, color=tertiaryColor, lw=2, ls='-', marker='*', label=molecule + ' VDE CORR')

    plt.xticks(l, labels, rotation = '30', ha='right')
    plt.margins(0.015, 0.05)
    plt.subplots_adjust(bottom=0.2, top=0.85)
    plt.ylabel(yLabel)
    plt.legend(loc='upper center', bbox_to_anchor=(.5, 1.2), numpoints = 1, shadow=True, ncol=3)
    plt.grid(True)
    if not os.path.exists(path+ '/ALL GRAPHS' + graphFolder+augmentedFolder):
            os.makedirs(path+ '/ALL GRAPHS' + graphFolder+augmentedFolder)
    plt.savefig(path+ '/ALL GRAPHS' + graphFolder+augmentedFolder + molecule + '.eps')
    #plt.show()
    plt.close()


def graphBothLinesVDE(valuesCation, valuesNeutral, l, labels, molecule, augmented, yLabel, graphFolder):
    if augmented==True:
        augmentedFolder=augFolder
    if augmented==False:
        augmentedFolder=ccFolder

    plt.plot(l, valuesCation, color=primaryColor, lw=2, ls='-', marker='s', label=molecule + ' cation')
    plt.plot(l, valuesNeutral, color=secondaryColor, lw=2, ls='-', marker='o', label=molecule + ' neutral')

    plt.xticks(l, labels, rotation = '30', ha='right')
    plt.margins(0.015, 0.05)
    plt.subplots_adjust(bottom=0.2, top=0.85)
    plt.ylabel(yLabel)
    plt.legend(loc='upper center', bbox_to_anchor=(.5, 1.2), numpoints = 1, shadow=True, ncol=2)
    plt.grid(True)

    if not os.path.exists(path + '/ALL GRAPHS' + graphFolder +augmentedFolder):
            os.makedirs(path + '/ALL GRAPHS' + graphFolder +augmentedFolder)
    plt.savefig(path + '/ALL GRAPHS' + graphFolder +augmentedFolder+ molecule + '.eps')
    #plt.show()
    plt.close()


def prepareGraph(startRow, endRow, molecule, worksheetName, augmented):

    #gets a list containing tuples of each of a molecules basis set, HF values and CORR
    #if worksheetName=="NEUTRAL":
    #    worksheetNumber=0
    #if worksheetName=="CHARGED":
    #    worksheetNumber=1
    #excelFile=pd.read_excel(path+'/logFiles.xlsx', sheetname=worksheetNumber)

    #worksheetName is either worksheetCharged or worksheetNeutral

    basisHFCORR=[]
    row=startRow

    print('row e' + str(row))
    print('end row' + str(endRow))

    while row<=endRow:
        hf=float(worksheetName[colHF + str(row)].value)
        corr=float(worksheetName[colCORR+str(row)].value)
        basis=str(worksheetName[colBasis+str(row)].value)
        ccsdt=float(worksheetName[colCCSDT+str(row)].value)
        tupBasis=(basis,hf,corr,ccsdt)
        basisHFCORR.append(tupBasis)
        print('tup basis yadada')
        print(tupBasis)
        row+=1
        #basisHFCORR has tuples with basis, hf, CORR, ccsdt
    ##print('Molecule is ' + molecule)
    ##print('Basis, hf, corr ) ' + str(basisHFCORR))

    #basisHFCORR is a list filled with tuples for just one molecule

    pvdz=None
    pvtz=None
    pvqz=None
    pv5z=None

    x=0
    ##print('len of basisHFDif')
    ##print(len(basisHFDif))
    print(basisHFCORR)
    while x < len(basisHFCORR):

        ##print('really long thing')
        ##print(basisHFDif[x][0][len(basisHFDif[x][0])-4:len(basisHFDif[x][0])])

        #checks if last four characters of basis (basisHFCORR[x][0]) matches with one in the basisSets
        if str(basisHFCORR[x][0][len(basisHFCORR[x][0])-4:len(basisHFCORR[x][0])])=='pVDZ':
            ##print('It')
            pvdz=basisHFCORR[x]
        if str(basisHFCORR[x][0][len(basisHFCORR[x][0])-4:len(basisHFCORR[x][0])])=='pVTZ':
            pvtz=basisHFCORR[x]
        if str(basisHFCORR[x][0][len(basisHFCORR[x][0])-4:len(basisHFCORR[x][0])])=='pVQZ':
            pvqz=basisHFCORR[x]
        if str(basisHFCORR[x][0][len(basisHFCORR[x][0])-4:len(basisHFCORR[x][0])])=='pV5Z':
            pv5z=basisHFCORR[x]
        x+=1

    ##print('pvdz')
    ##print(pvtz)
    ##print('pvtz')
    ##print(pvtz)
    ##print('pvqz')
    ##print(pvqz)
    ##print('pv5z')
    ##print(pv5z)

    labels=[]       #graph labels
    hfValues=[]     #y values
    corrValues=[]
    ccsdtValues=[]
    l=[]
    basisTuples=[pvdz, pvtz, pvqz, pv5z]
    x=0
    xAxis=0
    while x< len(basisTuples):
        ##print('basis Tuples [x]')
        ##print(basisTuples[x])
        if basisTuples[x] is not None:
            #print('hello basis tuples')
            labels.append(basisTuples[x][0])
            hfValues.append(basisTuples[x][1])
            corrValues.append(basisTuples[x][2])
            ccsdtValues.append(basisTuples[x][3])
            l.append(xAxis)
            xAxis+=5
        x+=1
    ##print(labels)
    ##print(hfValues)
    ##print(l)


    ylabelHF="HF Value"
    ylabelCORR="CCSD(T)-HF"
    ylabelHFandCCSDT='HF and CCSD(T) Values'
    ylabelHFandCORR='HF and CORR Values'
    graphFolderHF="/hf_graphs/"
    graphFolderCORR="/corr_graphs/"
    graphFolderHFandCCSDT='/HF_and_CCSD(T)/'
    graphFolderHFandCORR='/HF_and_CORR/'

    print(molecule)
    #print('Basis Tuples LIST')
    #print(basisTuples)


    '''COMMENTING OUT GRAPHS
    graph_OneLine(hfValues, l, labels, molecule, worksheetName, augmented, ylabelHF, graphFolderHF)
    graph_OneLine(corrValues, l, labels, molecule, worksheetName, augmented, ylabelCORR, graphFolderCORR)

    graph_TwoLines(hfValues, ccsdtValues, l, labels, molecule, worksheetName, augmented, ylabelHFandCCSDT, graphFolderHFandCCSDT, ' HF', ' CCSD(T)')
    graph_TwoLines(hfValues, corrValues, l, labels, molecule, worksheetName, augmented, ylabelHFandCORR, graphFolderHFandCORR, ' HF', ' CORR')
    '''

    """Example Basis Tuples List
    Basis Tuples LIST
[('CC-pVDZ', -398.6945222, -0.170250199999998, -398.8647724), ('CC-pVTZ', -398.7129813, -0.222371499999952, -398.9353528), ('CC-pVQZ', -398.7179851, -0.237516699999958, -398.9555018), ('CC-pV5Z', -398.7197233, -0.242467099999999, -398.9621904)]
Basis Tuples LIST
[('Aug-CC-pVDZ', -76.8275701, -0.293112899999997, -77.120683), ('Aug-CC-pVTZ', -76.8486209, -0.343532499999995, -77.1921534), ('Aug-CC-pVQZ', -76.8529819, -0.357889599999993, -77.2108715), ('Aug-CC-pV5Z', -76.8539425, -0.362375799999995, -77.2163183)]
    """

    findCORRvaluesForCorrSheet(molecule, augmented, worksheetName, basisTuples)

def prepareVDEgraph(startRow, endRow, molecule, augmented):
    '''gets a list containing tuples of a basis set, VDEhf, VDEccsdt, VDEcorr'''
    #excelFile=pd.read_excel(path+excelFilePathName, sheetname=sheetNumber)

    #use worksheetVDE automatically

    tupVDEarray=[]
    '''contains a list of tuples (basis set, VDEhf, VDEccsdt, VDEcorr, cat, HF neutral, CORR cat, Corr neutral)'''
    row=startRow
    while row<=endRow:

        basis=str(worksheetVDE[colVDEBasis+str(row)].value)
        VDE_HF=float(worksheetVDE[colVDE_HF_eV+str(row)].value)
        VDE_CCSDT=float(worksheetVDE[colVDE_CCSDT_eV+str(row)].value)
        VDE_CORR=float(worksheetVDE[colVDE_CORR_eV+str(row)].value)

        HF_cat=float(worksheetVDE[colVDE_HFcharged+str(row)].value)
        HF_neutral=float(worksheetVDE[colVDE_HFneutral+str(row)].value)

        CORR_cat=float(worksheetVDE[colVDE_CORRcharged+str(row)].value)
        CORR_neutral=float(worksheetVDE[colVDE_CORRneutral+str(row)].value)


        tup=(basis, VDE_HF, VDE_CCSDT, VDE_CORR, HF_cat, HF_neutral, CORR_cat, CORR_neutral)
        tupVDEarray.append(tup)
        row+=1


    pvdz=None
    pvtz=None
    pvqz=None
    pv5z=None

    x=0

    while x<len(tupVDEarray):
        if str(tupVDEarray[x][0][len(tupVDEarray[x][0])-4:len(tupVDEarray[x][0])])=='pVDZ':
            ##print('preparing VDE graph')
            pvdz=tupVDEarray[x]
        if str(tupVDEarray[x][0][len(tupVDEarray[x][0])-4:len(tupVDEarray[x][0])])=='pVTZ':
            pvtz=tupVDEarray[x]
        if str(tupVDEarray[x][0][len(tupVDEarray[x][0])-4:len(tupVDEarray[x][0])])=='pVQZ':
            pvqz=tupVDEarray[x]
        if str(tupVDEarray[x][0][len(tupVDEarray[x][0])-4:len(tupVDEarray[x][0])])=='pV5Z':
            pv5z=tupVDEarray[x]
        x+=1

    labels=[]       #graph labels
    VDE_HF_values=[]     #y values
    VDE_CCSDT_values=[]
    VDE_CORR_values=[]

    HF_cat_values=[]
    HF_neutral_values=[]
    CORR_cat_values=[]
    CORR_neutral_values=[]

    l=[]
    basisTuples=[pvdz, pvtz, pvqz, pv5z]
    x=0
    xAxis=0

    while x< len(basisTuples):
        if basisTuples[x] is not None:
            labels.append(basisTuples[x][0])
            VDE_HF_values.append(basisTuples[x][1])
            VDE_CCSDT_values.append(basisTuples[x][2])
            VDE_CORR_values.append(basisTuples[x][3])

            HF_cat_values.append(basisTuples[x][4])
            HF_neutral_values.append(basisTuples[x][5])
            CORR_cat_values.append(basisTuples[x][6])
            CORR_neutral_values.append(basisTuples[x][7])

            l.append(xAxis)
            xAxis+=5
        x+=1

    yLabelHF = 'VDE_HF=HF(N-1) - HF(N)'
    yLabelCCSDT = 'CCSDT_HF=CCSDT(N-1) - CCSDT(N)'
    yLabelCORR = 'CORR_HF=CORR(N-1) - CORR(N)'
    yLabelVDEcombined='VDE of HF, CCSD(T) and CORR'

    graphFolderHF = '/VDE_HF/'
    graphFolderCCSDT = '/VDE_CCSDT/'
    graphFolderCORR='/VDE_CORR/'
    graphFolderVDEcombined='/combinedVDE_CCSD(T)_HF_CORR/'



    graphVDE(VDE_HF_values, l, labels, molecule, augmented, yLabelHF, graphFolderHF)
    graphVDE(VDE_CCSDT_values, l, labels, molecule, augmented, yLabelCCSDT, graphFolderCCSDT)
    graphVDE(VDE_CORR_values, l, labels, molecule, augmented, yLabelCORR, graphFolderCORR)

    graph_VDEcombined_CCSDT_CORR_HF(VDE_HF_values, VDE_CCSDT_values, VDE_CORR_values, l, labels, molecule, worksheetVDE, augmented,\
    yLabelVDEcombined, graphFolderVDEcombined)


    yLabelBothHF='HF values for neutral and cation'
    graphFolderBothLinesHF='/HF_cat_and_neutral_bothLines/'

    yLabelBothCORR='CORR values for neutral and action'
    graphFolderBothLinesCORR='/CORR_cat_and_neutral_bothLines/'

    graphBothLinesVDE(HF_cat_values, HF_neutral_values, l, labels, molecule, augmented, yLabelBothHF, graphFolderBothLinesHF)
    graphBothLinesVDE(CORR_cat_values, CORR_neutral_values, l, labels, molecule, augmented, yLabelBothCORR, graphFolderBothLinesCORR)


def numberOfBasisSets(logarray):
    '''returns a list of the split log arrays by basis set. length is number of basis sets'''
    commandLocation=[]
    logsToReturn=[]
    x=0
    while x < len(logarray):
        if logarray[x] =='command:':
            commandLocation.append(x)
        x+=1
    commandLocation.append(len(logarray))
    ##print(commandLocation)
    x=0
    while x< len(commandLocation)-1:
        b=logarray[commandLocation[x]:commandLocation[x+1]]
        logsToReturn.append(b)
        x+=1
        ##print(b)
        ##print(len(b))
    ##print(logsToReturn)
    ##print(len(logsToReturn))
    return logsToReturn

def writeDataToExcel(worksheet, row, molecule, charge, multiplicity, basis, symmetry, hf, ccsdt,\
difference, mp2, mp3, mp4d, mp4dq, mp4sdq, ccsd, orbital, electronicState):
    '''function takes in values from dataExtract to add to excel file'''

    worksheet[colMolecule+str(row)]=molecule
    worksheet[colCharge+str(row)]=charge
    worksheet[colMultiplicity+str(row)]=multiplicity
    worksheet[colBasis+str(row)]=basis

    worksheet[colSym+str(row)]=symmetry
    worksheet[colHF+str(row)]=float(hf)
    worksheet[colCCSDT+str(row)]=float(ccsdt)
    worksheet[colCORR+str(row)]=difference

    worksheet[colMP2+str(row)]=float(mp2)
    worksheet[colMP3+str(row)]=float(mp3)
    worksheet[colMP4D+str(row)]=float(mp4d)
    worksheet[colMP4DQ+str(row)]=float(mp4dq)

    worksheet[colMP4SDQ+str(row)]=float(mp4sdq)
    worksheet[colCCSD+str(row)]=float(ccsd)
    worksheet[colOrbital+str(row)]=orbital
    worksheet[colElectronicState+str(row)]=electronicState

def dataExtract(path):
    '''prep excel workbook for NEUTRAL and CHARGED worksheets'''




    #add headings to each column in neutral worksheet
    worksheetNeutral[colMolecule+'1']='Molecule'
    worksheetNeutral[colCharge+'1']='Charge'
    worksheetNeutral[colBasis+'1']='Basis'
    worksheetNeutral[colCCSDT+'1']='CCSD(T)'
    worksheetNeutral[colHF+'1']='HF'
    worksheetNeutral[colCORR+'1']='CCSD(T)-HF'
    worksheetNeutral[colMP2+'1']='MP2'
    worksheetNeutral[colMP3+'1']='MP3'
    worksheetNeutral[colMP4D+'1']='MP4D'
    worksheetNeutral[colMP4DQ+'1']='MP4DQ'
    worksheetNeutral[colMP4SDQ+'1']='MP4SDQ'
    worksheetNeutral[colSym+'1']='Symmetry'
    worksheetNeutral[colMultiplicity+'1']='Multiplicity'
    worksheetNeutral[colCCSD+'1']='CCSD'
    worksheetNeutral[colOrbital+'1']='Orbital'
    worksheetNeutral[colElectronicState+'1']='Electronic State'

    #add headings to each column in charged worksheet
    worksheetCharged[colMolecule+'1']='Molecule'
    worksheetCharged[colCharge+'1']='Charge'
    worksheetCharged[colBasis+'1']='Basis'
    worksheetCharged[colCCSDT+'1']='CCSD(T)'
    worksheetCharged[colHF+'1']='HF'
    worksheetCharged[colCORR+'1']='CCSD(T)-HF'
    worksheetCharged[colMP2+'1']='MP2'
    worksheetCharged[colMP3+'1']='MP3'
    worksheetCharged[colMP4D+'1']='MP4D'
    worksheetCharged[colMP4DQ+'1']='MP4DQ'
    worksheetCharged[colMP4SDQ+'1']='MP4SDQ'
    worksheetCharged[colSym+'1']='Symmetry'
    worksheetCharged[colMultiplicity+'1']='Multiplicity'
    worksheetCharged[colCCSD+'1']='CCSD'
    worksheetCharged[colOrbital+'1']='Orbital'
    worksheetCharged[colElectronicState+'1']='Electronic State'

    '''function extracts data from log files and sends to writeDataToExcel'''

    logFiles=[]

    for path, subdirs, files in os.walk(path+logFilesFolder):
        for name in files:
            if os.path.join(path, name)[len(os.path.join(path, name))-4:len(os.path.join(path, name))]=='.log':
                logFiles.append(os.path.join(path, name))

    #logFiles=['/Users/Jared/Dropbox/Auburn/Research/Second_Research/Log_Files/CH4_Dz.txt']
    #,'CH4_Qz.txt','CH4_Tz.txt'
    #list of all the log files to open. might need to change to a loop later

    molAndSymNeutral=[]   #holds list of tuples of all the molecules/symmetry combos. To use in function returnRowsforSymsandMols
    molAndSymCharged=[]

    rowNeutral = 2
    rowCharged = 2
    ##print(logFiles)
    '''begin searching and saving data'''
    for currentFile in logFiles:
        log = open(currentFile, 'r').read()

        splitLog = re.split(r'[\\\s]\s*', log)  #splits string with \ (\\), empty space (\s) and = and ,
        ###print(repr(splitlog))
        ###print(splitlog)
        ###print(len(splitlog))
        #open up the log file, read it, and split it- log
        ###print('length of splitlog is' + str(len(splitlog)))
        #reset all values

        for splitlog in numberOfBasisSets(splitLog): #NUMBEROFSPLITS will return where in log file it needs to be split for basis sets
            #textFile(log)   #text file will return each log split by basis set because some aren't
            ##print(splitlog)
            molecule = None
            charge = None
            multiplicity = None
            basis = None
            symmetry=None
            hf = None
            ccsdt = None
            difference = None
            mp2=None
            mp3=None
            mp4d=None
            mp4dq=None
            mp4sdq=None
            ccsd=None
            orbital=None
            electronicState=None
            valuesBlockFound=False
            populationAnalysisFound=False    #need for orbital
            alphaOrbitalFound=False     #need for multiplicity not 1 for electronic state

            x=0
            while x<len(splitlog):

                                                    #find name of molecule
                if splitlog[x] == 'Stoichiometry':
                    moleculeWithCharge = splitlog[x+1]
                    molecule=''
                    for w in moleculeWithCharge:
                        if w!='(':
                            molecule+=w
                        else:
                            break
                                                #find charge and multiplicity
                if splitlog[x] == 'Multiplicity':
                    charge = float(splitlog[x-1])
                    ##print('charge is ' + str(charge))
                    multiplicity = int(splitlog[x+2])
                    ##print('multiplicity is ' + str(multiplicity))
                                                #find basis and symmetry
                if splitlog[x] == 'Standard' and splitlog[x+1]=='basis:':
                    basis = splitlog[x+2]
                    ##print('basis is ' + basis)

                if splitlog[x]=='Full':
                    symmetry=splitlog[x+3]
                    ##print('symmetry is ' + symmetry)
                                                    #find block with data

                if splitlog[x]=='Population' and splitlog[x+1]=='analysis':
                    #print(multiplicity)
                    populationAnalysisFound=True
                    ##print('population analysis found')

                if multiplicity==1 and populationAnalysisFound==True and splitlog[x]=='Virtual' and orbital==None:
                    orbital = splitlog[x-1]
                    ##print('orbital found for multiplicity is 1')

                if splitlog[x]=='Alpha' and splitlog[x+1]=='Orbitals:' and populationAnalysisFound==True and multiplicity!=None and multiplicity!=1:
                    alphaOrbitalFound=True
                    ##print('alpha orbital found')

                if alphaOrbitalFound==True and splitlog[x]=='Virtual' and orbital==None:
                    orbital=splitlog[x-1]
                    ##print('orbital found for multiplicity not 1')

                if populationAnalysisFound==True and electronicState==None:
                    if splitlog[x]=='The' or splitlog[x]=='Unable':
                        y=0
                        while splitlog[x+y]!='Alpha':
                            y+=1
                        electronicState=''.join(splitlog[x:x+y])



                if splitlog[x]=='SP':
                    valuesBlockFound=True

                    y=0
                    while splitlog[x+y]!='@':
                        y+=1
                    valuesBlock=''.join(splitlog[x:x+y])
                    ##print('start values block')
                    ##print(valuesBlock)
                    ##print('end values block')
                    l=0
                    while l < len(valuesBlock):
                                                        #find HF
                        if valuesBlock[l:l+3]=='HF=' and valuesBlock[l-2:l]!='PU':
                            start=l+3
                            end=l+5
                            numberDone=False
                            while numberDone==False:
                                end+=1
                                try:
                                    float(valuesBlock[start:end])
                                except:
                                    numberDone=True
                            hf=valuesBlock[start:end-1]
                            ##print('HF is ' + hf)
                                                            #find CCSD(T)
                                                            #calculate difference
                        if valuesBlock[l:l+8]=='CCSD(T)=':
                            start=l+8
                            end=l+11
                            numberDone=False
                            while numberDone==False:
                                end+=1
                                try:
                                    float(valuesBlock[start:end])
                                except:
                                    numberDone=True
                            ccsdt=valuesBlock[start:end-1]
                            ##print('CCSD(T) is ' + ccsdt)

                            difference = float(ccsdt) - float(hf)
                                                            #find MP2
                        if valuesBlock[l:l+4]=='MP2=':
                            start=l+4
                            end=l+6
                            numberDone=False
                            while numberDone==False:
                                end+=1
                                try:
                                    float(valuesBlock[start:end])
                                except:
                                    numberDone=True
                            mp2=valuesBlock[start:end-1]
                            ##print('MP2 is ' + mp2)
                                                        #find MP3
                        if valuesBlock[l:l+4]=='MP3=':
                            start=l+4
                            end=l+6
                            numberDone=False
                            while numberDone==False:
                                end+=1
                                try:
                                    float(valuesBlock[start:end])
                                except:
                                    numberDone=True
                            mp3=valuesBlock[start:end-1]
                            ##print('MP3 is ' + mp3)
                                                        #find MP4D
                        if valuesBlock[l:l+5]=='MP4D=':
                            start=l+5
                            end=l+6
                            numberDone=False
                            while numberDone==False:
                                end+=1
                                try:
                                    float(valuesBlock[start:end])
                                except:
                                    numberDone=True
                            mp4d=valuesBlock[start:end-1]
                            ##print('MP4D is ' + mp4d)
                                                        #find MP4DQ
                        if valuesBlock[l:l+6]=='MP4DQ=':
                            start=l+6
                            end=l+7
                            numberDone=False
                            while numberDone==False:
                                end+=1
                                try:
                                    float(valuesBlock[start:end])
                                except:
                                    numberDone=True
                            mp4dq=valuesBlock[start:end-1]
                            ##print('MP4DQ is ' + mp4dq)
                                                            #find MP4SDQ
                        if valuesBlock[l:l+7]=='MP4SDQ=':
                            start=l+7
                            end=l+8
                            numberDone=False
                            while numberDone==False:
                                end+=1
                                try:
                                    float(valuesBlock[start:end])
                                except:
                                    numberDone=True
                            mp4sdq=valuesBlock[start:end-1]
                            ##print('MP4SDQ is ' + mp4sdq)
                                                            #find CCSD
                        if valuesBlock[l:l+5]=='CCSD=':
                            start=l+5
                            end=l+6
                            numberDone=False
                            while numberDone==False:
                                end+=1
                                try:
                                    float(valuesBlock[start:end])
                                except:
                                    numberDone=True
                            ccsd=valuesBlock[start:end-1]
                            ##print('CCSD is ' + ccsd)

                        l+=1
                x+=1

            if valuesBlockFound==True:
                ##print('molecule is ' + molecule)
                ##print('HF is ' + hf)


                '''figure out which worksheet to use'''
                ##print('Charge is ' + str(charge))
                molsym=(str(molecule), str(symmetry), str(basis[0:2]))

                if charge==0:
                    worksheet=worksheetNeutral
                    row = rowNeutral
                    if molsym not in molAndSymNeutral:
                        molAndSymNeutral.append(molsym)
                    #molAndSym=molAndSymNeutral
                else:
                    worksheet=worksheetCharged
                    row = rowCharged
                    if molsym not in molAndSymCharged:
                        molAndSymCharged.append(molsym)

                #data stored in variables is input to writesDataToExcel
                writeDataToExcel(worksheet, row, molecule, charge, multiplicity, basis, symmetry, hf, ccsdt,\
        difference, mp2, mp3, mp4d, mp4dq, mp4sdq, ccsd, orbital, electronicState)

                if charge==0:
                    rowNeutral+=1
                else:
                    rowCharged+=1

    VDEexcel(worksheetCharged, worksheetNeutral)
    CreateCORRexcel()

    ##print(molAndSymNeutral)
    for m in molAndSymNeutral:
        print('m is ')
        print(m)
        #molAndSym contains tuples of (molecule, symmetry)
        #m[0] gives molecule
        #m[1] gives symmetry
        #m[2] gives basis
        startEnd=rowsForSymAndMol(m[0], m[1], m[2], worksheetNeutral)
        print('rowsforsymandmol input is ')
        print(m[0], m[1], m[2], worksheetNeutral)
        print('worksheet neutral E5')
        print(worksheetNeutral['E5'].value)
        startRow=startEnd[0]
        endRow=startEnd[1]
        print('neutral end row ' + str(endRow))
        if m[2][0:2]=='Au':             #augmented is a boolean. true if augmented. false if not.
            augmented=True
        elif m[2][0:2]=='CC':
            augmented=False
        ##print(m[0], m[2][0:2], startRow+2, endRow+2)
        #if endRow-startRow>=4:
         #   #print('MORE THAN FOUR ')
          #  #print(m[0], m[1], m[2], 'NEUTRAL')
        prepareGraph(startRow, endRow, m[0], worksheetNeutral, augmented)
        prepareVDEgraph(startRow, endRow, m[0], augmented)

    ##print(molAndSymCharged)
    for m in molAndSymCharged:
        #molAndSym contains tuples of (molecule, symmetry)
        #m[0] gives molecule
        #m[1] gives symmetry
        #m[2] gives basis. [0:2] is either cc or au
        startEnd=rowsForSymAndMol(m[0], m[1], m[2], worksheetCharged)
        startRow=startEnd[0]
        endRow=startEnd[1]
        print('charged end row ' + str(endRow))
        if m[2][0:2]=='Au':
            augmented=True
        elif m[2][0:2]=='CC':
            augmented=False
        ##print(m[0], m[2][0:2], startRow+2, endRow+2)
        #if endRow-startRow>=4:
         #   #print('MORE THAN FOUR ')
          #  #print(m[0], m[1], m[2], 'CHARGED')
        prepareGraph(startRow, endRow, m[0], worksheetCharged, augmented)


        #df = DataFrame({'Molecule':arrMolecule})
        #df.to_excel('/Users/Jared/Dropbox/Auburn/Research/Second Research/log Files/logFiles.xlsx', sheet_name='sheet1',index = False)



    workbook.save(pathorigin + excelFilePathName)     #saves file

def VDEexcel(chargedWorksheet, neutralWorksheet):
    '''function creates an excel sheet for VDE difference between charged and neutral'''



    worksheetVDE[colVDEMolecule+'1']='Molecule'
    worksheetVDE[colVDESymmetry+'1']='Symmetry'
    worksheetVDE[colVDEBasis+'1']='Basis'
    worksheetVDE[colVDE_HFcharged+'1']='HF (N-1)'
    worksheetVDE[colVDE_HFneutral+'1']='HF (N)'
    worksheetVDE[colVDE_HF_Ha+'1']='VDE(HF)(Ha)=HF(N-1)-HF(N)'
    worksheetVDE[colVDE_HF_eV+'1']='VDE(HF) (eV)'
    worksheetVDE[colVDE_CCSDTcharged+'1']='CCSD(T) (N-1)'
    worksheetVDE[colVDE_CCSDTneutral+'1']='CCSD(T) (N)'
    worksheetVDE[colVDE_CCSDT_Ha+'1']='VDE(CCSDT)(Ha)=CCSD(T)(N-1)-CCSD(T)(N)'
    worksheetVDE[colVDE_CCSDT_eV+'1']='VDE(CCSDT)(eV)=CCSD(T)(N-1)-CCSD(T)(N)'
    worksheetVDE[colVDE_CORRcharged+'1']='CORR (N-1)'
    worksheetVDE[colVDE_CORRneutral+'1']='CORR (N)'
    worksheetVDE[colVDE_CORR_Ha+'1']='VDE(CORR)(Ha)=CORR(N-1) - CORR(N)'
    worksheetVDE[colVDE_CORR_eV+'1']='VDE(CORR)(eV)=CORR(N-1) - CORR(N)'

    row=2
    totalRows=neutralWorksheet.max_row

    while row<=totalRows:

        '''set all VDE worksheet values to none'''
        neutralMolecule=None
        chargedMolecule=None
        neutralSymmetry=None
        chargedSymmetry=None
        neutralBasis=None
        chargedBasis=None
        neutralHF=None
        chargedHF=None
        VDE_HF=None
        neutralCCSDT=None
        chargedCCSDT=None
        VDE_CCSDT=None
        neutralCORR=None
        chargedCORR=None
        VDE_CORR=None

        '''find all the necessary values from neutral worksheet'''
        neutralMolecule=neutralWorksheet[colMolecule + str(row)].value
        neutralSymmetry=neutralWorksheet[colSym + str(row)].value
        neutralBasis=neutralWorksheet[colBasis + str(row)].value

        neutralHF=neutralWorksheet[colHF + str(row)].value
        neutralCCSDT=neutralWorksheet[colCCSDT + str(row)].value
        neutralCORR=neutralWorksheet[colCORR + str(row)].value

        '''write neutral values into VDE worksheet'''
        worksheetVDE[colVDEMolecule+str(row)]=neutralMolecule
        worksheetVDE[colVDESymmetry+str(row)]=neutralSymmetry
        worksheetVDE[colVDEBasis+str(row)]=neutralBasis

        worksheetVDE[colVDE_HFneutral+str(row)]=neutralHF
        worksheetVDE[colVDE_CCSDTneutral+str(row)]=neutralCCSDT
        worksheetVDE[colVDE_CORRneutral+str(row)]=neutralCORR


        '''find all necessary values from charged worksheet. Since not all the charged and neutral molecules are lined
        up properly between the worksheets a loop is run to find where the charged worksheet molecule matches the
        neutral worksheet molecule'''
        otherRow=2
        while otherRow<=totalRows:
            """'''find what the charged row molecule symmetry and basis is'''"""
            chargedMolecule=chargedWorksheet[colMolecule + str(otherRow)].value
            chargedSymmetry=chargedWorksheet[colSym + str(otherRow)].value
            chargedBasis=chargedWorksheet[colBasis + str(otherRow)].value

            '''check if charged values == neutral values to make sure'''
            if chargedMolecule==neutralMolecule and chargedSymmetry==neutralSymmetry \
            and chargedBasis==neutralBasis:
                ''' if they are the same proceed and find the hf ccsdt and corr values'''
                chargedHF=chargedWorksheet[colHF + str(otherRow)].value
                chargedCCSDT=chargedWorksheet[colCCSDT + str(otherRow)].value
                chargedCORR=chargedWorksheet[colCORR + str(otherRow)].value
                '''write charged values into VDE worksheet'''

                worksheetVDE[colVDE_HFcharged+str(otherRow)]=chargedHF
                worksheetVDE[colVDE_CCSDTcharged+str(otherRow)]=chargedCCSDT
                worksheetVDE[colVDE_CORRcharged+str(otherRow)]=chargedCORR
                break   #since the charged and neutral rows match up break from loop
            otherRow+=1


        '''equate all VDE values and write into excel'''
        VDE_HF=chargedHF-neutralHF
        VDE_CCSDT=chargedCCSDT-neutralCCSDT
        VDE_CORR=chargedCORR-neutralCORR

        if chargedBasis[0]=='A':
            augmented=True
        elif chargedBasis[0]=='C':
            augmented=False

        #vdeCCSDTev_list[chargedMolecule, chargedBasis]=VDE_CCSDT*Ha_eV_conversion
        x=0
        added=False
        while x<len(vdeCCSDTev_list):
            if vdeCCSDTev_list[x][0]==chargedMolecule and vdeCCSDTev_list[x][1]==augmented:
                vdeCCSDTev_list[x].append([chargedBasis, VDE_CCSDT*Ha_eV_conversion])
                added=True
            x+=1
        if added==False:
            vdeCCSDTev_list.append([chargedMolecule, augmented, [chargedBasis, VDE_CCSDT*Ha_eV_conversion]])
            added=True
        #stores VDE CCSD(T) values in a dictionary. Key is molecule and basis

        #graphBothLinesVDE(chargedHF, neutralHF, l, labels, molecule, augmented, yLabel, graphFolder):
        #creat graphs for HF N, N-1
        #create graphs for CCSDT N, N-1
        #create graphs for CORR N, N-1

        worksheetVDE[colVDE_HF_Ha+str(row)]=VDE_HF
        worksheetVDE[colVDE_CCSDT_Ha+str(row)]=VDE_CCSDT
        worksheetVDE[colVDE_CORR_Ha+str(row)]=VDE_CORR

        worksheetVDE[colVDE_HF_eV+str(row)]=VDE_HF*Ha_eV_conversion
        worksheetVDE[colVDE_CCSDT_eV+str(row)]=VDE_CCSDT*Ha_eV_conversion
        worksheetVDE[colVDE_CORR_eV+str(row)]=VDE_CORR*Ha_eV_conversion

        row+=1

def CreateCORRexcel():


    worksheetCORR[colCORRmolecule+'1']='Molecule'
    worksheetCORR[colCORRaugmented+'1']='Augmented'
    worksheetCORR[colCORRbasisX+'1']='Basis Set X'

    worksheetCORR[colCORRbasisY+'1']='Basis Set Y'
    worksheetCORR[colCORRcharge+'1']='Charge'
    worksheetCORR[colCORR_Ecorr+'1']='Ecorr xy'

    worksheetCORR[colCORR_A+'1']='A'
    worksheetCORR[colCORR_VDE_CORR+'1']='CORR VDE'
    worksheetCORR[colMaxBasisSet+'1']='Largest Basis HF(x)'

    worksheetCORR[colVDEcc_hfcorr+'1']='ccVDE hf+corr'
    worksheetCORR[colVDEcc_eV+'1']='ccVDE (eV)'

def findCORRvaluesForCorrSheet(molecule, augmented, worksheetName, moleculeTuple):
    #print(molecule, augmented, worksheetName, moleculeTuple)
    ''' example tuple input
    labels.append(basisTuples[x][0])
            hfValues.append(basisTuples[x][1])
            corrValues.append(basisTuples[x][2])
            ccsdtValues.append(basisTuples[x][3])
    [(basis, hf, corr, ccsdt)...]
    Basis Tuples LIST
[('CC-pVDZ', -398.6945222, -0.170250199999998, -398.8647724), ('CC-pVTZ', -398.7129813, -0.222371499999952, -398.9353528),
('CC-pVQZ', -398.7179851, -0.237516699999958, -398.9555018), ('CC-pV5Z', -398.7197233, -0.242467099999999, -398.9621904)]
Basis Tuples LIST
[('Aug-CC-pVDZ', -76.8275701, -0.293112899999997, -77.120683), ('Aug-CC-pVTZ', -76.8486209, -0.343532499999995, -77.1921534), ('Aug-CC-pVQZ', -76.8529819, -0.357889599999993, -77.2108715), ('Aug-CC-pV5Z', -76.8539425, -0.362375799999995, -77.2163183)]
    '''
    #need to write to excel (molecule, augmented, basis set x, basis set y

    global corrSheetRow


    maxHF=None

    n=0
    while maxHF==None and n<len(moleculeTuple)-1:
        if moleculeTuple[len(moleculeTuple)-1-n]!=None:
            maxHF=moleculeTuple[len(moleculeTuple)-1-n][1]
            maxBasisSet=moleculeTuple[len(moleculeTuple)-1-n][0]
        n+=1

    basisXcounter=0
    basisYcounter=1
    while basisXcounter<len(moleculeTuple)-1:
        while basisYcounter<len(moleculeTuple):
            if basisYcounter>basisXcounter and moleculeTuple[basisYcounter]!=None and moleculeTuple[basisXcounter]!=None:
                print('you are now here')
                ''' if moleculeTuple[0][0][0]=='C':
                    augmented=False
                elif moleculeTuple[0][0][0]=='A':
                    augmented=True
                '''
                basisSetX=moleculeTuple[basisXcounter][0][len(moleculeTuple[basisXcounter][0])-4:len(moleculeTuple[basisXcounter][0])]
                basisSetY=moleculeTuple[basisYcounter][0][len(moleculeTuple[basisYcounter][0])-4:len(moleculeTuple[basisYcounter][0])]

                basisSetXFull=moleculeTuple[basisXcounter][0]
                basisSetYFull=moleculeTuple[basisYcounter][0]
                #PVDZ, PVTZ, etc.


                xCorrValue=moleculeTuple[basisXcounter][2]
                yCorrValue=moleculeTuple[basisYcounter][2]

                eCORRxy = CalculateEcorrXY(xCorrValue, basisSetX, yCorrValue, basisSetY)
                a_value=CalculateA(eCORRxy, xCorrValue, basisSetX)


                #worksheet[colMP2+str(row)]=float(mp2)

                if worksheetName==worksheetCharged:
                    worksheetName='CHARGED'
                elif worksheetName==worksheetNeutral:
                    worksheetName='NEUTRAL'

                worksheetCORR[colCORRmolecule+str(corrSheetRow)]=molecule
                worksheetCORR[colCORRaugmented+str(corrSheetRow)]=augmented
                worksheetCORR[colCORRbasisX+str(corrSheetRow)]=basisSetX

                worksheetCORR[colCORRbasisY+str(corrSheetRow)]=basisSetY
                worksheetCORR[colCORRcharge+str(corrSheetRow)]=worksheetName
                worksheetCORR[colCORR_Ecorr+str(corrSheetRow)]=eCORRxy
                worksheetCORR[colCORR_A+str(corrSheetRow)]=a_value

                '''
                worksheetCORR.write(corrSheetRow, colCORRmolecule, molecule)
                worksheetCORR.write(corrSheetRow, colCORRaugmented, augmented)
                worksheetCORR.write(corrSheetRow, colCORRbasisX, basisSetX)

                worksheetCORR.write(corrSheetRow, colCORRbasisY, basisSetY)
                worksheetCORR.write(corrSheetRow, colCORRcharge, worksheetName)
                worksheetCORR.write(corrSheetRow, colCORR_Ecorr, eCORRxy)
                worksheetCORR.write(corrSheetRow, colCORR_A, a)
                '''

                #correlationMoleculeDict={}
                #takes the molecule and basis sets and saves the row and eCORRxy
                #correlationMoleculeDict[x]=4
                if worksheetName==neutralName:
                    correlationMoleculeDict[(molecule, augmented, basisSetX, basisSetY)]=(corrSheetRow, eCORRxy, maxHF)
                elif worksheetName==chargedName:
                    corrCharged=eCORRxy
                    corrNeutral=correlationMoleculeDict[(molecule, augmented, basisSetX, basisSetY)][1]
                    corrVDErow=correlationMoleculeDict[(molecule, augmented, basisSetX, basisSetY)][0]

                    corrVDE=corrCharged-corrNeutral
                    #worksheetCORR.write(corrVDErow, colCORR_VDE_CORR, corrVDE)
                    worksheetCORR[colCORR_VDE_CORR+str(corrVDErow)]=corrVDE


                    maxHFcharged=maxHF
                    maxHFneutral=correlationMoleculeDict[(molecule, augmented, basisSetX,basisSetY)][2]

                    ccVDE_hf_corr=(maxHFcharged+corrCharged)-(maxHFneutral+corrNeutral)


                    worksheetCORR[colMaxBasisSet+str(corrVDErow)]=maxBasisSet
                    worksheetCORR[colVDEcc_hfcorr+str(corrVDErow)]=ccVDE_hf_corr
                    worksheetCORR[colVDEcc_eV+str(corrVDErow)]=ccVDE_hf_corr*Ha_eV_conversion
                    '''
                    worksheetCORR.write(corrVDErow, colMaxBasisSet, maxBasisSet)
                    worksheetCORR.write(corrVDErow, colVDEcc_hfcorr, ccVDE_hf_corr)
                    worksheetCORR.write(corrVDErow, colVDEcc_eV, ccVDE_hf_corr*Ha_eV_conversion)
                    '''

                    #vdeCCSDTeV_dict[(molecule,)]=(ccVDE_hf_corr*Ha_eV_conversion, basisSetX, basisSetY)
                    #vdeCCSDTlist.append((molecule, augmented)

                    x=0
                    while x<len(vdeCCSDTev_list):
                        ##print('BBBBBBBB')
                        ##print(molecule, augmented)
                        if vdeCCSDTev_list[x][0]==molecule and vdeCCSDTev_list[x][1]==augmented:
                            vdeCCSDTev_list[x].append([basisSetXFull, basisSetYFull, ccVDE_hf_corr*Ha_eV_conversion])
                        x+=1

                corrSheetRow+=1

            basisYcounter+=1
        basisXcounter+=1
        basisYcounter=basisXcounter+1

    #vdeCCSDTeV_list[chargedMolecule, chargedBasis]=VDE_CCSDT*Ha_eV_conversion

    global timesRun
    timesRun+=1
    #print(timesRun)


    if timesRun==88:
        #print('vdeCCSDTev_list')
        #print vdeCCSDTev_list
        for x in vdeCCSDTev_list:
            graphVDEcorr_indBasis_and_Extrapolated(x)


def CalculateEcorrXY(xCorr, xBasis, yCorr, yBasis):

    basisNumberDict={'D':2,'T':3,'Q':4,'5':5}
    x=basisNumberDict[xBasis[2]]
    y=basisNumberDict[yBasis[2]]

    numerator=(xCorr*x**3)-(yCorr*y**3)
    denominator=(x**3)-(y**3)
    eCORRxy=numerator/denominator
    return eCORRxy

def CalculateA(eCORRxy, eCORRx, basis):
    #inputs x can also be y

    basisNumberDict={'D':2,'T':3,'Q':4,'5':5}
    x=basisNumberDict[basis[2]]


    n=eCORRx-eCORRxy
    m=x**3
    a=n*m
    return a

def graphVDEcorr_indBasis_and_Extrapolated(vdeCCSDT):
    #print('VDE CCSDT LIST GIVEN')
    #print(vdeCCSDT)

    global timesRun
    timesRun+=1
    #print(timesRun)

    molecule=vdeCCSDT[0]
    augmented=vdeCCSDT[1]

    labels=[]
    values=[]
    l=[]

    basisAndCCSDTvalues=vdeCCSDT[2:]

    xAxis=0
    x=0
    while x< len(basisAndCCSDTvalues):


        values.append(basisAndCCSDTvalues[x][len(basisAndCCSDTvalues[x])-1])
        thisLabel=''
        y=0
        while y<len(basisAndCCSDTvalues[x])-1:
            thisLabel+=basisAndCCSDTvalues[x][y] + ' '
            y+=1
        labels.append(thisLabel)

        l.append(xAxis)
        xAxis+=5

        x+=1

    graphVDE(values, l, labels, molecule, augmented, 'VDE CCSD(T) (eV)', '/Compare CCSDT values/')
    #values, l-5,10,15... , basis labels, molecule, augmented-T/F, yLabel string, graphFolder)
